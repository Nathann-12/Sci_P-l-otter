"""
Tests for file_io module.
"""

import os
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Callable, Iterable, List
from unittest.mock import MagicMock, patch

import numpy as np
import pandas as pd
import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from file_io import read_file, read_csv, read_excel, read_cdf


class TestFileIO:
    def setup_method(self):
        self.sample_data = pd.DataFrame({
            "time": pd.date_range("2024-01-01", periods=100, freq="1h"),
            "value": np.random.randn(100),
            "category": ["A", "B"] * 50,
        })
        self.small_df = self.sample_data.head(6).copy()
        self.temp_dir = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    # -------------------------------
    # helpers
    # -------------------------------
    def _df_to_rows(self, df: pd.DataFrame) -> List[List[object]]:
        rows: List[List[object]] = []
        for row in df.itertuples(index=False):
            converted: List[object] = []
            for value in row:
                if isinstance(value, pd.Timestamp):
                    converted.append(value.to_pydatetime())
                else:
                    converted.append(value)
            rows.append(converted)
        return rows

    def _create_workbook(self, name: str, builder: Callable[[Workbook], None]) -> Path:
        path = Path(self.temp_dir) / name
        wb = Workbook()
        builder(wb)
        wb.save(path)
        return path

    def _write_table(self, ws, start_row: int, start_col: int, headers: Iterable[str], rows: Iterable[Iterable[object]]):
        for idx, header in enumerate(headers):
            ws.cell(row=start_row, column=start_col + idx, value=header)
        for offset, values in enumerate(rows, start=1):
            for idx, value in enumerate(values):
                ws.cell(row=start_row + offset, column=start_col + idx, value=value)
        return ws

    # -------------------------------
    # CSV tests
    # -------------------------------
    def test_read_csv_basic(self):
        csv_path = Path(self.temp_dir) / "test.csv"
        self.sample_data.to_csv(csv_path, index=False)

        df, metadata = read_csv(csv_path)

        assert df is not None
        assert not df.empty
        assert len(df) == len(self.sample_data)
        assert metadata["source"] == "csv"

    def test_read_csv_with_different_delimiters(self):
        csv_path = Path(self.temp_dir) / "test_semicolon.csv"
        self.sample_data.to_csv(csv_path, sep=";", index=False)

        df, _ = read_csv(csv_path)
        assert not df.empty

        tsv_path = Path(self.temp_dir) / "test_tab.tsv"
        self.sample_data.to_csv(tsv_path, sep="\t", index=False)

        df, _ = read_csv(tsv_path)
        assert not df.empty

    def test_read_csv_encoding_detection(self):
        csv_path = Path(self.temp_dir) / "encoded.csv"
        self.sample_data.to_csv(csv_path, index=False, encoding="utf-8")

        df, metadata = read_csv(csv_path)
        assert not df.empty
        assert metadata["encoding"] in {"utf-8", "utf-8-sig"}

    def test_read_csv_empty_file(self):
        csv_path = Path(self.temp_dir) / "empty.csv"
        csv_path.write_text("", encoding="utf-8")

        with pytest.raises(pd.errors.EmptyDataError):
            read_csv(csv_path)

    # -------------------------------
    # Excel tests
    # -------------------------------
    def test_read_excel_basic(self):
        rows = self._df_to_rows(self.small_df)
        excel_path = self._create_workbook(
            "basic.xlsx",
            lambda wb: self._write_table(wb.active, 1, 1, list(self.small_df.columns), rows),
        )

        df, metadata = read_excel(excel_path)

        assert list(df.columns) == list(self.small_df.columns)
        assert len(df) == len(self.small_df)
        assert metadata["source"] == "excel"
        assert metadata["tables"]
        assert metadata["tables"][0]["rows"] == len(df)

    def test_read_excel_offset_detection(self):
        rows = self._df_to_rows(self.small_df)

        def builder(wb):
            ws = wb.active
            ws.title = "OffsetData"
            ws["A1"] = "Summary"
            ws["A2"] = "Generated for tests"
            self._write_table(ws, 5, 2, list(self.small_df.columns), rows)

        excel_path = self._create_workbook("offset.xlsx", builder)

        df, metadata = read_excel(excel_path)

        assert list(df.columns) == list(self.small_df.columns)
        assert len(df) == len(self.small_df)
        assert metadata["sheet"] == "OffsetData"
        assert metadata["tables"][0]["header_row"] == 5
        assert metadata["tables"][0]["cols"] == len(self.small_df.columns)

    def test_read_excel_multiple_tables(self):
        primary_rows = self._df_to_rows(self.small_df)
        secondary_rows = [["Segment A", 10], ["Segment B", 20], ["Segment C", 30]]

        def builder(wb):
            ws = wb.active
            ws.title = "Multi"
            self._write_table(ws, 3, 2, list(self.small_df.columns), primary_rows)
            self._write_table(ws, 12, 7, ["category", "score"], secondary_rows)

        excel_path = self._create_workbook("tables.xlsx", builder)

        df_primary, meta_primary = read_excel(excel_path)
        assert not df_primary.empty
        assert len(meta_primary["tables"]) >= 2

        tables = meta_primary["tables"]
        secondary_idx = next(
            idx for idx, table in enumerate(tables)
            if table["cols"] == 2 and table["rows"] == len(secondary_rows)
        )

        df_secondary, meta_secondary = read_excel(excel_path, table_index=secondary_idx)
        assert len(df_secondary) == len(secondary_rows)
        assert set(df_secondary.columns) == {"category", "score"}
        assert meta_secondary["table_index"] == secondary_idx

    def test_read_excel_sheet_selection_and_range(self):
        primary_rows = self._df_to_rows(self.small_df)

        def builder(wb):
            ws1 = wb.active
            ws1.title = "Overview"
            self._write_table(ws1, 2, 1, ["id", "name"], [[1, "alpha"], [2, "beta"]])
            ws2 = wb.create_sheet("DeepData")
            self._write_table(ws2, 6, 4, list(self.small_df.columns), primary_rows)

        excel_path = self._create_workbook("sheets.xlsx", builder)

        df_sheet, meta_sheet = read_excel(excel_path, sheet="DeepData")
        assert not df_sheet.empty
        assert meta_sheet["sheet"] == "DeepData"
        assert meta_sheet["tables"][0]["header_row"] == 6

        df_range, meta_range = read_excel(excel_path, sheet="DeepData", data_range="D6:F12")
        assert len(df_range) == len(self.small_df)
        assert meta_range["tables"][0]["origin"] == "range"

    def test_read_excel_header_override(self):
        rows = self._df_to_rows(self.small_df)

        def builder(wb):
            ws = wb.active
            ws.title = "Headers"
            ws.cell(row=4, column=2, value="Group")
            ws.cell(row=4, column=3, value="Group")
            self._write_table(ws, 5, 2, list(self.small_df.columns), rows)

        excel_path = self._create_workbook("override.xlsx", builder)

        _, meta_default = read_excel(excel_path)
        assert meta_default["tables"][0]["header_row"] == 5

        df_override, meta_override = read_excel(excel_path, header_row=4)
        assert not df_override.empty
        assert meta_override["tables"][0]["header_row"] == 4
        assert any(col.startswith("Group") for col in meta_override["tables"][0]["columns"])

    def test_read_excel_repeated_header_row_removed(self):
        headers = ["name", "value"]
        rows = [
            ["alpha", "10"],
            headers,  # repeated header row
            ["beta", "20"],
        ]

        def builder(wb):
            ws = wb.active
            ws.title = "DupHeaders"
            self._write_table(ws, 4, 3, headers, rows)

        excel_path = self._create_workbook("dup_headers.xlsx", builder)

        df, _ = read_excel(excel_path)
        assert len(df) == 2
        assert df.iloc[0]["value"] == 10.0
        assert df.iloc[1]["value"] == 20.0

    def test_read_excel_type_coercion(self):
        headers = ["value", "timestamp", "flag", "note"]
        rows = [
            ["1", "2024-01-01 00:00:00", "TRUE", " text "],
            ["2.5", "2024-01-02", "FALSE", ""],
            [None, None, "TRUE", "   "],
        ]

        def builder(wb):
            ws = wb.active
            ws.title = "Coerce"
            ws["A1"] = "Metadata"
            self._write_table(ws, 3, 2, headers, rows)

        excel_path = self._create_workbook("coerce.xlsx", builder)

        df, _ = read_excel(excel_path)
        assert pd.api.types.is_numeric_dtype(df["value"])
        assert pd.api.types.is_datetime64_any_dtype(df["timestamp"])
        assert df["flag"].astype(str).str.upper().isin({"TRUE", "FALSE"}).all()
        assert df["note"].iloc[0] == "text"
        assert pd.isna(df["value"].iloc[-1])

    def test_read_excel_table_index_out_of_range(self):
        rows = self._df_to_rows(self.small_df)
        excel_path = self._create_workbook(
            "single_table.xlsx",
            lambda wb: self._write_table(wb.active, 2, 2, list(self.small_df.columns), rows),
        )

        with pytest.raises(ValueError, match="table_index"):
            read_excel(excel_path, table_index=10)

    # -------------------------------
    # read_file integration
    # -------------------------------
    def test_read_file_auto_detection(self):
        csv_path = Path(self.temp_dir) / "auto.csv"
        self.sample_data.to_csv(csv_path, index=False)

        df_csv, meta_csv = read_file(csv_path)
        assert not df_csv.empty
        assert meta_csv["source"] == "csv"

        rows = self._df_to_rows(self.small_df)
        excel_path = self._create_workbook(
            "auto.xlsx",
            lambda wb: self._write_table(wb.active, 1, 1, list(self.small_df.columns), rows),
        )

        df_xlsx, meta_xlsx = read_file(excel_path)
        assert not df_xlsx.empty
        assert meta_xlsx["source"] == "excel"

    def test_read_file_nonexistent(self):
        with pytest.raises(FileNotFoundError):
            read_file(Path(self.temp_dir) / "missing.csv")

    def test_read_file_unsupported_extension(self):
        path = Path(self.temp_dir) / "note.xyz"
        path.write_text("dummy", encoding="utf-8")
        with pytest.raises(ValueError):
            read_file(path)

    # -------------------------------
    # CDF / NetCDF (mocked)
    # -------------------------------
    @patch("file_io.read_netcdf_quick")
    def test_read_netcdf_mock(self, mock_quick):
        sample = self.sample_data.head(3)
        mock_quick.return_value = (sample, {"source": "netcdf", "path": "mock"})

        nc_path = Path(self.temp_dir) / "test.nc"
        nc_path.write_text("dummy", encoding="utf-8")

        df, metadata = read_file(nc_path)
        assert df.equals(sample)
        assert metadata["source"] == "netcdf"


if __name__ == "__main__":
    pytest.main([__file__])
